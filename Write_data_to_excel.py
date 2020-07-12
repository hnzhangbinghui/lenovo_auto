import xlrd
import xlwt
import pandas as pd
import openpyxl
from xlutils.copy import copy
value=1000
file_path=r"C:\Users\Administrator\Desktop\write_excel.xlsx"
# workbook=openpyxl.Workbook()
# sheet=workbook.active
# sheet.title='Sheet'
# for i in range(1,10):
#     sheet.cell(row=i,column=5,value=value)
# workbook.save(file_path)

#对excel表内容进行追加
data=openpyxl.load_workbook(file_path)
sheetnames=data.get_sheet_names()
table=data.get_sheet_by_name(sheetnames[0])
table=data.active
print(table.title)
nrows=table.max_row
ncols=table.max_column
r=2
c=4
values=['加油2020']*10
for v in values:
    table.cell(r+1,4).value=v
    r=r+1
data.save(file_path)






# workbook=xlrd.open_workbook(,formatting_info=True)
# booksheet=workbook.sheet_by_index(0)
# rows=booksheet.nrows
# cols=booksheet.ncols
# print("行数：" ,rows)
# print("列数：",cols)
# sheet=workbook.get_sheet(0)
# sheet.write(1,1,1000)









#打开目标表格
# wb=openpyxl.load_workbook(r"C:\Users\Administrator\Desktop\write_excel.xlsx")
# #打开sheet
# ws=wb['Sheet1']
# distance_list=[]
# ws.cell(row=2,column=8).value='1000'
# # for i in range(1,len(distance_list)+1):
# #     distance=distance_list[i-1]
# #     # 写入位置的行列号可以任意改变，这里我是从第2行开始按行依次插入第11列
# #     ws.cell(row = i+1, column = 11).value =distance
# wb.save(r"C:\Users\Administrator\Desktop\write_excel.xlsx")





# eight_values=booksheet.col_values(8,1,rows)
# print(eight_values)